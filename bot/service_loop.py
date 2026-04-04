import os
import shutil
import time

import numpy as np
import openpyxl
import pyautogui
from PIL import Image

import config
from utils import wait_and_click, wait_for_image
from schedule_capture import capture_schedule
from uploader import upload_service, get_existing_services
from tsw_api import get_player_location
from navigator import (
    pass_warning_screen,
    pass_splash_screen,
    click_to_the_trains,
    click_choose_a_route,
    select_route,
    click_timetable,
    select_extra_section,
    select_train,
    click_train,
    exit_game,
    relaunch_and_navigate,
)


class ServiceError(Exception):
    """Raised when a service fails, carrying position info for resume."""

    def __init__(self, train_name, train_number, service_number, original_error=None):
        self.train_name = train_name
        self.train_number = train_number        # 1-indexed
        self.service_number = service_number    # 1-indexed
        super().__init__(
            f"Failed at {train_name} train {train_number} service {service_number}: "
            f"{original_error}"
        )


class BatchRestart(Exception):
    """Raised when batch limit reached, carrying resume position."""

    def __init__(self, train_name, train_number, service_number):
        self.train_name = train_name
        self.train_number = train_number        # 1-indexed
        self.service_number = service_number    # 1-indexed
        super().__init__(
            f"Batch limit reached at {train_name} train {train_number}, "
            f"resuming at service {service_number}"
        )


def get_spreadsheet_path(train_name):
    """Build the spreadsheet path for a train.

    Converts train name to filename: '/' → '_', lowercase, spaces → '_'.
    Checks for .xlsx first, then .ods.
    E.g. 'Class 350/1' → 'class_350_1.xlsx' or 'class_350_1.ods'
    """
    base = train_name.lower().replace("/", "_").replace(" ", "_")
    directory = os.path.join(config.SPREADSHEETS_DIR, config.ROUTE_NAME)

    for ext in (".xlsx", ".ods"):
        path = os.path.join(directory, base + ext)
        if os.path.isfile(path):
            return path

    # Default to .xlsx for the error message
    return os.path.join(directory, base + ".xlsx")


def _read_ods_spreadsheet(path):
    """Read service names, bounds, and first stations from an ODS spreadsheet.

    Returns a list of dicts:
        [{"name": "741-06", "bound": "northbound", "first_station": "Kilburn High Road"}, ...]
    """
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = load(path)
    services = []

    for table in doc.spreadsheet.getElementsByType(Table):
        sheet_name = table.getAttribute("name")
        bound = sheet_name.strip().lower()

        # Build a 2D list of cell values
        rows = []
        for row in table.getElementsByType(TableRow):
            cells = []
            for cell in row.getElementsByType(TableCell):
                # Handle repeated columns
                repeat = cell.getAttribute("numbercolumnsrepeated")
                repeat = int(repeat) if repeat else 1
                # Get cell text content
                text_parts = []
                for p in cell.getElementsByType(P):
                    # Get all text content from the paragraph
                    text = ""
                    for child in p.childNodes:
                        if hasattr(child, 'data'):
                            text += child.data
                        elif hasattr(child, '__str__'):
                            text += str(child)
                    text_parts.append(text)
                value = "\n".join(text_parts) if text_parts else None
                # Cap repeats to avoid huge empty expansions
                for _ in range(min(repeat, 256)):
                    cells.append(value)
            rows.append(cells)

        if not rows:
            continue

        # Find the actual number of columns (from row 1, the header row)
        max_col = len(rows[0]) if rows else 0

        # Each column (index 1 onward, i.e. B+) is one service; row 0 has the service name
        for col in range(1, max_col):
            svc_name = rows[0][col] if col < len(rows[0]) else None
            if not svc_name or not str(svc_name).strip():
                continue
            svc_name = str(svc_name).strip()

            # Find first station: first non-None, non-'-' value from row 2 onward
            first_station = None
            for row_idx in range(2, len(rows)):
                val = rows[row_idx][col] if col < len(rows[row_idx]) else None
                if val is not None and str(val).strip() not in ("-", ""):
                    station = rows[row_idx][0] if len(rows[row_idx]) > 0 else None
                    first_station = str(station).strip() if station else None
                    break

            services.append({
                "name": svc_name,
                "bound": bound,
                "first_station": first_station,
            })

    return services


def read_spreadsheet(train_name):
    """Read service names, bounds, and first stations from a spreadsheet.

    Supports both .xlsx (openpyxl) and .ods (odfpy) formats.

    Returns a list of dicts:
        [{"name": "741-06", "bound": "northbound", "first_station": "Kilburn High Road"}, ...]
    """
    path = get_spreadsheet_path(train_name)
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Spreadsheet not found: {path}")

    if path.endswith(".ods"):
        return _read_ods_spreadsheet(path)

    wb = openpyxl.load_workbook(path, data_only=True)
    services = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        bound = sheet_name.strip().lower()  # "Northbound" → "northbound"

        # Each column (B onward) is one service; row 1 has the service name
        for col in range(2, ws.max_column + 1):
            svc_name = ws.cell(1, col).value
            if not svc_name:
                continue
            svc_name = str(svc_name).strip()

            # Find first station: first non-None, non-'-' value from row 3 onward
            first_station = None
            for row in range(3, ws.max_row + 1):
                val = ws.cell(row, col).value
                if val is not None and str(val).strip() not in ("-", ""):
                    first_station = str(ws.cell(row, 1).value).strip()
                    break

            services.append({
                "name": svc_name,
                "bound": bound,
                "first_station": first_station,
            })

    wb.close()
    return services


def search_and_select_service(service_name):
    """Type a service name into the search field to find and select it.

    1. Click the service name search field (service_name_search.png)
    2. Clear existing text
    3. Type the service name
    4. Wait for the list to filter
    5. Click the first result (top box)
    6. Wait for highlight confirmation

    Returns (success, y_center) where y_center is the highlighted box position.
    """
    # Click the search field
    if not wait_and_click(config.REF_SERVICE_NAME_SEARCH,
                          timeout=config.SCREEN_TIMEOUT,
                          confidence=config.CONFIDENCE, verify=False):
        print(f"       ERROR: Could not find service search field")
        return False, None
    time.sleep(0.5)

    # Clear existing text and type the service name
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.2)
    pyautogui.typewrite(service_name, interval=0.03)
    time.sleep(2.0)  # wait for filter to update

    # Click the first visible service box (top of the list)
    boxes = get_visible_service_boxes()
    if not boxes:
        print(f"       ERROR: No service boxes visible after search")
        return False, None

    x, y = boxes[0]
    pyautogui.moveTo(x, y)
    time.sleep(0.5)
    pyautogui.mouseDown()
    time.sleep(0.2)
    pyautogui.mouseUp()
    time.sleep(1.5)

    # Confirm highlight
    highlight_count, actual_y = wait_for_single_highlight()
    if highlight_count == 1:
        return True, actual_y if actual_y else y
    elif highlight_count > 1:
        # Multiple matches — still use the first one
        print(f"       WARNING: {highlight_count} boxes highlighted, using first")
        return True, actual_y if actual_y else y
    else:
        print(f"       WARNING: No highlight after clicking search result")
        return False, None


def clear_service_search():
    """Clear the service search field to show all services again."""
    found = wait_and_click(config.REF_SERVICE_NAME_SEARCH,
                           timeout=5,
                           confidence=config.CONFIDENCE, verify=False)
    if not found:
        # Search icon not found — click the known search field area as fallback
        print(f"       Search icon not found, using fallback clear")
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.1)
    pyautogui.press("delete")
    time.sleep(1.0)


def get_visible_service_boxes():
    """Calculate the positions of all 8 visible service boxes using fixed stride.

    Uses FIRST_BOX_TOP offset from SERVICE_LIST_TOP, then SERVICE_BOX_STRIDE
    for each subsequent box. Returns list of (center_x, center_y) tuples in
    screen coordinates.
    """
    center_x = (config.SERVICE_LIST_LEFT + config.SERVICE_LIST_RIGHT) // 2
    boxes = []
    for i in range(config.BOXES_PER_PAGE):
        box_top = config.FIRST_BOX_TOP + i * config.SERVICE_BOX_STRIDE
        box_center_y = config.SERVICE_LIST_TOP + box_top + config.SERVICE_BOX_HEIGHT // 2
        if box_top + config.SERVICE_BOX_HEIGHT > config.SERVICE_LIST_BOTTOM - config.SERVICE_LIST_TOP:
            break
        boxes.append((center_x, box_center_y))
    return boxes



def wait_for_single_highlight():
    """Wait until exactly one service box has the #57a6d0 border.

    Scans the full service list left edge for the border color.
    Retries a few times if zero or multiple boxes are highlighted.
    Returns (count, center_y) where center_y is the screen Y of the
    highlighted box (or None if count != 1).
    """
    target = np.array(config.SERVICE_BOX_BORDER_RGB)
    tol = config.SERVICE_BOX_COLOR_TOLERANCE
    min_height = 30

    for attempt in range(2):
        region = (
            config.SERVICE_LIST_LEFT,
            config.SERVICE_LIST_TOP,
            config.SERVICE_LIST_RIGHT - config.SERVICE_LIST_LEFT,
            config.SERVICE_LIST_BOTTOM - config.SERVICE_LIST_TOP,
        )
        img = np.array(pyautogui.screenshot(region=region))

        strip = img[:, :5, :].astype(int)
        match = np.all(np.abs(strip - target) <= tol, axis=2)
        row_has_border = np.any(match, axis=1)

        runs = []
        in_run = False
        run_start = 0
        for r in range(len(row_has_border)):
            if row_has_border[r] and not in_run:
                run_start = r
                in_run = True
            elif not row_has_border[r] and in_run:
                runs.append((run_start, r))
                in_run = False
        if in_run:
            runs.append((run_start, len(row_has_border)))

        valid_runs = [(s, e) for s, e in runs if (e - s) >= min_height]
        box_count = len(valid_runs)

        if box_count == 1:
            s, e = valid_runs[0]
            center_y = config.SERVICE_LIST_TOP + (s + e) // 2
            print(f"       Single highlight confirmed (attempt {attempt + 1})")
            return 1, center_y

        print(f"       Highlight check: found {box_count} bordered boxes (attempt {attempt + 1}), waiting...")
        time.sleep(0.5)

    print(f"       WARNING: Expected 1 highlighted box, found {box_count}")
    return box_count, None


def _grab_and_crop_service_box(y_center, padding=30):
    """Grab an oversized region around a service box and crop to the #57a6d0 border.

    Returns (cropped_img, raw_img, found_border) where found_border is True if
    the blue border was detected and used for cropping. raw_img is the full
    uncropped grab for debugging.
    """
    grab_left = config.SERVICE_LIST_LEFT
    grab_top = y_center - config.SERVICE_BOX_HEIGHT // 2 - padding
    grab_width = config.SERVICE_LIST_RIGHT - config.SERVICE_LIST_LEFT + config.BOX_EXTRA_RIGHT
    grab_height = config.SERVICE_BOX_HEIGHT + 2 * padding

    screenshot = pyautogui.screenshot(region=(grab_left, grab_top, grab_width, grab_height))
    img = np.array(screenshot)  # RGB
    raw_img = img.copy()

    target = np.array(config.SERVICE_BOX_BORDER_RGB)
    tol = config.SERVICE_BOX_COLOR_TOLERANCE
    strip = img[:, :5, :].astype(int)
    match = np.all(np.abs(strip - target) <= tol, axis=2)
    row_match = np.any(match, axis=1)

    runs = []
    in_run = False
    run_start = 0
    for r in range(len(row_match)):
        if row_match[r] and not in_run:
            run_start = r
            in_run = True
        elif not row_match[r] and in_run:
            runs.append((run_start, r))
            in_run = False
    if in_run:
        runs.append((run_start, len(row_match)))

    if not runs:
        return img, raw_img, False

    expected_center = padding + config.SERVICE_BOX_HEIGHT // 2
    best_run = min(runs, key=lambda r: abs((r[0] + r[1]) / 2 - expected_center))
    row_top, row_bottom = best_run

    box_strip = img[row_top:row_bottom, :, :].astype(int)
    col_match = np.all(np.abs(box_strip - target) <= tol, axis=2)
    col_has_border = np.any(col_match, axis=0)

    left = 0
    right = img.shape[1]
    border_cols = np.where(col_has_border)[0]
    if len(border_cols) > 0:
        left = int(border_cols[0])
        right = int(border_cols[-1]) + 1

    img = img[row_top:row_bottom, left:right, :]
    return img, raw_img, True


def _validate_border(img):
    """Check that the cropped image has a solid #57a6d0 border on all 4 edges.

    Returns (has_blue, has_edge_border) where:
      has_blue: image contains any #57a6d0 pixels at all
      has_edge_border: all 4 edges have a continuous 2px blue border (>=90% blue)
                       AND image height is at least 80% of SERVICE_BOX_HEIGHT
    """
    target = np.array(config.SERVICE_BOX_BORDER_RGB)
    tol = config.SERVICE_BOX_COLOR_TOLERANCE
    h, w = img.shape[:2]
    min_px = 2
    min_ratio = 0.90  # at least 90% of edge pixels must be blue

    all_match = np.all(np.abs(img.astype(int) - target) <= tol, axis=2)
    has_blue = bool(np.any(all_match))

    if not has_blue:
        return False, False

    # Reject if the crop is too short — likely a partial box
    min_height = int(config.SERVICE_BOX_HEIGHT * 0.8)
    if h < min_height:
        print(f"       Border validation: image too short ({h}px < {min_height}px min)")
        return True, False

    edges_ok = True
    edge_names = ["top", "bottom", "left", "right"]
    edges = [img[:min_px, :, :],       # top
             img[-min_px:, :, :],       # bottom
             img[:, :min_px, :],        # left
             img[:, -min_px:, :]]       # right

    for name, edge in zip(edge_names, edges):
        match = np.all(np.abs(edge.astype(int) - target) <= tol, axis=-1)
        ratio = np.mean(match)
        if ratio < min_ratio:
            print(f"       Border validation: {name} edge only {ratio:.0%} blue (need {min_ratio:.0%})")
            edges_ok = False
            break

    return has_blue, edges_ok


def _find_highlighted_box_y():
    """Scan the service list for the currently highlighted box and return its y_center.

    Looks for the #57a6d0 border in the full service list area.
    Returns the screen y_center of the highlighted box, or None if not found.
    """
    region = (config.SERVICE_LIST_LEFT, config.SERVICE_LIST_TOP,
              config.SERVICE_LIST_RIGHT - config.SERVICE_LIST_LEFT,
              config.SERVICE_LIST_BOTTOM - config.SERVICE_LIST_TOP)
    img = np.array(pyautogui.screenshot(region=region))
    target = np.array(config.SERVICE_BOX_BORDER_RGB)
    tol = config.SERVICE_BOX_COLOR_TOLERANCE

    strip = img[:, :5, :].astype(int)
    match = np.all(np.abs(strip - target) <= tol, axis=2)
    row_has_border = np.any(match, axis=1)

    runs = []
    in_run = False
    run_start = 0
    for r in range(len(row_has_border)):
        if row_has_border[r] and not in_run:
            run_start = r
            in_run = True
        elif not row_has_border[r] and in_run:
            runs.append((run_start, r))
            in_run = False
    if in_run:
        runs.append((run_start, len(row_has_border)))

    boxes = [(s, e) for s, e in runs if (e - s) >= 30]
    if boxes:
        s, e = boxes[0]
        return config.SERVICE_LIST_TOP + (s + e) // 2
    return None


def screenshot_service_box(output_dir, y_center, click_x=None, click_y=None):
    """Take a validated screenshot of a service box.

    Crops to the #57a6d0 border on all 4 sides, then validates:
    1. Blue border present — if not, retake screenshot
    2. Still no blue — re-click the box and retake
    3. Still no blue — save for debugging and raise error
    4. Blue present but not on all edges — retake with larger grab
    5. Still clipped — save with warning

    Saves as 1_service.png inside output_dir.
    """
    path = os.path.join(output_dir, "1_service.png")

    # Attempt 1: normal grab
    img, raw, found = _grab_and_crop_service_box(y_center)
    if found:
        has_blue, has_edges = _validate_border(img)
    else:
        has_blue, has_edges = False, False

    if has_blue and has_edges:
        Image.fromarray(img).save(path)
        return path

    # Attempt 2: retake screenshot
    if not has_blue:
        print("       No blue border in screenshot, retaking...")
        time.sleep(0.5)
        img, raw, found = _grab_and_crop_service_box(y_center)
        if found:
            has_blue, has_edges = _validate_border(img)
        else:
            has_blue, has_edges = False, False

        if has_blue and has_edges:
            Image.fromarray(img).save(path)
            return path

    # Attempt 3: re-click + retake
    if not has_blue:
        if click_x is not None and click_y is not None:
            print("       Still no blue border, re-clicking box and retaking...")
            pyautogui.moveTo(click_x, click_y)
            time.sleep(0.5)
            pyautogui.mouseDown()
            time.sleep(0.2)
            pyautogui.mouseUp()
            time.sleep(1.5)
            img, raw, found = _grab_and_crop_service_box(y_center)

            if found:
                has_blue, has_edges = _validate_border(img)
            else:
                has_blue, has_edges = False, False

            if has_blue and has_edges:
                Image.fromarray(img).save(path)
                return path

    # No blue at all after 3 attempts — save for debugging and raise
    if not has_blue:
        Image.fromarray(img).save(path)
        raise RuntimeError(
            f"No blue border (#57a6d0) detected in service screenshot after 3 attempts"
        )

    # Blue is present but edges are clipped — the box is partially off-screen
    # (e.g. at the bottom of the scroll container). Micro-scroll to bring it
    # into view and retake. No need to scroll back — the loop exits to main
    # menu after each service and re-navigates from the top.
    if not has_edges:
        target = np.array(config.SERVICE_BOX_BORDER_RGB)
        tol = config.SERVICE_BOX_COLOR_TOLERANCE
        min_ratio = 0.90
        bottom_match = np.all(np.abs(img[-2:, :, :].astype(int) - target) <= tol, axis=-1)
        top_match = np.all(np.abs(img[:2, :, :].astype(int) - target) <= tol, axis=-1)
        bottom_ok = float(np.mean(bottom_match)) >= min_ratio
        top_ok = float(np.mean(top_match)) >= min_ratio
        h = img.shape[0]
        height_ok = h >= int(config.SERVICE_BOX_HEIGHT * 0.8)

        if not bottom_ok or not top_ok or not height_ok:
            # Determine scroll direction: if bottom is worse or image is too short,
            # scroll up (positive) to push content down; otherwise scroll down
            if not bottom_ok or (not height_ok and top_ok):
                direction = "bottom"
                scroll_nudge = config.SCROLL_PER_BOX
            else:
                direction = "top"
                scroll_nudge = -config.SCROLL_PER_BOX
            print(f"       Border clipped on {direction} "
                  f"(h={h}px, top={np.mean(top_match):.0%}, bottom={np.mean(bottom_match):.0%}), "
                  f"micro-scrolling to bring box into view...")

            center_x = (config.SERVICE_LIST_LEFT + config.SERVICE_LIST_RIGHT) // 2
            center_y = (config.SERVICE_LIST_TOP + config.SERVICE_LIST_BOTTOM) // 2
            pyautogui.moveTo(center_x, center_y)
            time.sleep(0.3)
            pyautogui.scroll(scroll_nudge)
            time.sleep(1.0)

            # Find where the highlighted box actually is now
            new_y = _find_highlighted_box_y()
            if new_y is not None:
                img, raw, found = _grab_and_crop_service_box(new_y)
    
                if found:
                    _, has_edges = _validate_border(img)

        if not has_edges:
            print("       WARNING: Border still clipped after scroll adjustment, saving anyway")

    Image.fromarray(img).save(path)
    return path


def click_service_box(x, y):
    """Click on a service box, then press Enter twice to load the level."""
    pyautogui.moveTo(x, y)
    time.sleep(0.5)
    pyautogui.mouseDown()
    time.sleep(0.2)
    pyautogui.mouseUp()
    time.sleep(1.5)           # give game time to highlight/select the box
    pyautogui.press("enter")
    time.sleep(1.0)
    pyautogui.press("enter")


def _check_for_level_screen():
    """Check once for driver selection or get_started screen.

    Returns 'driver' if driver screen found (and clicks it),
    'get_started' if get_started screen found, or None if neither.
    """
    # Try each driver reference image
    for driver_ref in (config.REF_DRIVER_1, config.REF_DRIVER_2):
        if not os.path.isfile(driver_ref):
            continue
        try:
            driver_loc = pyautogui.locateOnScreen(driver_ref, confidence=config.CONFIDENCE)
            if driver_loc is not None:
                print(f"       Driver selection detected via '{os.path.basename(driver_ref)}' — clicking...")
                center = pyautogui.center(driver_loc)
                pyautogui.moveTo(center)
                time.sleep(0.5)
                pyautogui.mouseDown()
                time.sleep(0.2)
                pyautogui.mouseUp()
                time.sleep(config.CLICK_SETTLE_DELAY)
                return "driver"
        except pyautogui.ImageNotFoundException:
            pass

    # Try each get_started reference image
    for started_ref in (config.REF_GET_STARTED_1, config.REF_GET_STARTED_2):
        if not os.path.isfile(started_ref):
            continue
        try:
            started_loc = pyautogui.locateOnScreen(started_ref, confidence=config.CONFIDENCE)
            if started_loc is not None:
                return "get_started"
        except pyautogui.ImageNotFoundException:
            pass

    return None


def wait_for_level_load(click_x=None, click_y=None, service_dir=None):
    """Wait for the level to load, handle optional driver selection, then get past 'Get Started'.

    If click coordinates are provided, retries the service box click up to RETRY_MAX times
    (waiting RETRY_WAIT seconds between attempts) if the level doesn't load.
    If service_dir is provided, captures a screenshot of the level info area before clicking
    'Get Started'.
    """
    print("       Waiting for level to load...")

    for attempt in range(1, config.RETRY_MAX + 1):
        # Wait RETRY_WAIT seconds for either driver or get_started screen
        start = time.time()
        found = None
        while time.time() - start < config.RETRY_WAIT:
            found = _check_for_level_screen()
            if found is not None:
                break
            time.sleep(1.0)

        if found is not None:
            break

        # Screen hasn't changed — retry the click if we have coordinates
        if attempt < config.RETRY_MAX:
            if click_x is not None and click_y is not None:
                print(f"       Screen hasn't changed (attempt {attempt}/{config.RETRY_MAX}), "
                      f"clicking service at ({click_x}, {click_y}) again...")
                click_service_box(click_x, click_y)
            else:
                print(f"       Screen hasn't changed (attempt {attempt}/{config.RETRY_MAX}), "
                      f"waiting again...")
        else:
            raise TimeoutError("Timed out waiting for level to load after "
                               f"{config.RETRY_MAX} attempts")

    # Now wait for either get_started variant and click it
    print("       Waiting for 'Get Started' screen...")
    get_started_refs = [r for r in (config.REF_GET_STARTED_1, config.REF_GET_STARTED_2)
                        if os.path.isfile(r)]
    found_loc = None
    start = time.time()
    while time.time() - start < config.SCREEN_TIMEOUT:
        for ref in get_started_refs:
            try:
                loc = pyautogui.locateOnScreen(ref, confidence=config.CONFIDENCE)
                if loc is not None:
                    found_loc = loc
                    break
            except pyautogui.ImageNotFoundException:
                pass
        if found_loc is not None:
            break
        time.sleep(1.0)
    if found_loc is None:
        raise TimeoutError("Timed out waiting for 'Get Started' screen")

    # Click 'Get Started' with retry — re-click if it's still visible
    for click_attempt in range(1, config.RETRY_MAX + 1):
        print("       Level loaded — clicking 'Get Started'...")
        center = pyautogui.center(found_loc)
        pyautogui.moveTo(center)
        time.sleep(0.5)
        pyautogui.mouseDown()
        time.sleep(0.2)
        pyautogui.mouseUp()
        time.sleep(3.0)

        # Check if Get Started is still on screen
        still_visible = False
        for ref in get_started_refs:
            try:
                loc = pyautogui.locateOnScreen(ref, confidence=config.CONFIDENCE)
                if loc is not None:
                    still_visible = True
                    found_loc = loc
                    break
            except pyautogui.ImageNotFoundException:
                pass

        if not still_visible:
            break

        if click_attempt < config.RETRY_MAX:
            print(f"       'Get Started' still visible (attempt {click_attempt}/{config.RETRY_MAX}), "
                  f"clicking again...")
        else:
            print(f"       WARNING: 'Get Started' still visible after {config.RETRY_MAX} attempts")

    time.sleep(5.0)           # game needs time to transition into gameplay


def exit_to_main_menu():
    """Press Escape, find 'Exit to Main Menu', and press Enter twice."""
    print("       Pressing Escape...")
    pyautogui.press("escape")
    time.sleep(4.0)           # pause menu needs time to render

    print("       Clicking 'Exit to Main Menu'...")
    main_menu_refs = [r for r in [config.REF_EXIT_TO_MAIN_MENU_1, config.REF_EXIT_TO_MAIN_MENU_2]
                      if os.path.isfile(r)]
    if not main_menu_refs:
        raise FileNotFoundError("No 'Exit to Main Menu' reference images found")
    clicked = False
    start = time.time()
    while time.time() - start < config.SCREEN_TIMEOUT:
        for ref in main_menu_refs:
            try:
                location = pyautogui.locateOnScreen(ref, confidence=config.CONFIDENCE,
                                                            region=config.GAME_REGION)
            except pyautogui.ImageNotFoundException:
                location = None
            if location is not None:
                center = pyautogui.center(location)
                pyautogui.moveTo(center)
                time.sleep(0.5)
                pyautogui.mouseDown()
                time.sleep(0.2)
                pyautogui.mouseUp()
                clicked = True
                break
        if clicked:
            break
        time.sleep(1.0)
    if not clicked:
        raise TimeoutError("Timed out waiting for 'Exit to Main Menu'")
    time.sleep(1.0)
    print("       Pressing Enter twice...")
    pyautogui.press("enter")
    time.sleep(0.5)
    pyautogui.press("enter")
    time.sleep(config.CLICK_SETTLE_DELAY)


def _count_visible_trains(img):
    """Count visible train entries by scanning the left edge for #dedede borders.

    If no trains are detected (e.g. the only train is highlighted in #4a90b5
    instead of the normal #dedede), falls back to checking for the highlight
    color and returns 1.

    Returns the number of distinct train entries visible in the image.
    """
    border_rgb = np.array([0xde, 0xde, 0xde])
    highlight_rgb = np.array([0x4a, 0x90, 0xb5])
    border_tol = 20
    min_entry_height = 40

    strip = img[:, :3, :].astype(int)
    match = np.all(np.abs(strip - border_rgb) <= border_tol, axis=2)
    row_has_border = np.any(match, axis=1)

    runs = []
    in_run = False
    run_start = 0
    for r in range(len(row_has_border)):
        if row_has_border[r] and not in_run:
            run_start = r
            in_run = True
        elif not row_has_border[r] and in_run:
            runs.append((run_start, r))
            in_run = False
    if in_run:
        runs.append((run_start, len(row_has_border)))

    count = len([(s, e) for s, e in runs if (e - s) >= min_entry_height])

    # Also check for a highlighted/selected train (#4a90b5) — this train
    # won't have the normal #dedede border, so count it separately.
    match_hl = np.all(np.abs(strip - highlight_rgb) <= border_tol, axis=2)
    row_has_hl = np.any(match_hl, axis=1)
    hl_runs = []
    in_hl = False
    hl_start = 0
    for r in range(len(row_has_hl)):
        if row_has_hl[r] and not in_hl:
            hl_start = r
            in_hl = True
        elif not row_has_hl[r] and in_hl:
            hl_runs.append((hl_start, r))
            in_hl = False
    if in_hl:
        hl_runs.append((hl_start, len(row_has_hl)))
    hl_count = len([(s, e) for s, e in hl_runs if (e - s) >= min_entry_height])
    count += hl_count

    return count


def count_trains():
    """Count the number of trains in the current train scroll box.

    Detects visible trains from the first frame, then scrolls one box at
    a time until the list stops moving. Returns visible + scrolls.
    Scrolls back to the top when done.
    """
    print("       Counting trains...")

    def _capture():
        region = (config.TRAIN_BOX_LEFT, config.TRAIN_BOX_TOP,
                  config.TRAIN_BOX_WIDTH, config.TRAIN_BOX_HEIGHT)
        return np.array(pyautogui.screenshot(region=region))

    def _frames_match(a, b):
        if a.shape != b.shape:
            return False
        return np.mean(np.abs(a.astype(float) - b.astype(float))) < 5.0

    first_img = _capture()
    visible_count = _count_visible_trains(first_img)
    print(f"       Visible trains in first frame: {visible_count}")

    # If all trains fit on screen (fewer than the max visible), no scrolling needed
    if visible_count < config.TRAIN_VISIBLE_COUNT:
        print(f"       All {visible_count} trains visible (< {config.TRAIN_VISIBLE_COUNT}), no scroll needed.")
        return visible_count

    prev_img = first_img
    scrolls = 0
    center_x = config.TRAIN_BOX_LEFT + config.TRAIN_BOX_WIDTH // 2
    center_y = config.TRAIN_BOX_TOP + config.TRAIN_BOX_HEIGHT // 2

    for _ in range(30):
        pyautogui.moveTo(center_x, center_y)
        time.sleep(0.3)
        pyautogui.scroll(config.TRAIN_SCROLL_PER_BOX)
        time.sleep(1.0)

        curr_img = _capture()
        if _frames_match(prev_img, curr_img):
            break
        scrolls += 1
        prev_img = curr_img

    total = visible_count + scrolls

    # Scroll back to top
    if scrolls > 0:
        pyautogui.moveTo(center_x, center_y)
        time.sleep(0.3)
        pyautogui.scroll(-config.TRAIN_SCROLL_PER_BOX * scrolls)
        time.sleep(1.0)

    print(f"       Found {total} trains ({visible_count} visible + {scrolls} scrolls).")
    return total


def _return_to_main_menu_from_menus():
    """From any in-game menu screen, press Escape until we reach the main menu.

    Checks for the 'To The Trains' tile to confirm we're at the main menu.
    Used when we're at the service list or class selection and need to get
    back to the main menu (not from inside a level — use exit_to_main_menu for that).
    """
    for attempt in range(6):
        # Check if we're already at the main menu (try both normal and hovered refs)
        at_menu = False
        for ref in [config.REF_TO_THE_TRAINS_1, config.REF_TO_THE_TRAINS_2]:
            if not os.path.isfile(ref):
                continue
            try:
                loc = pyautogui.locateOnScreen(ref, confidence=config.CONFIDENCE)
                if loc is not None:
                    at_menu = True
                    break
            except pyautogui.ImageNotFoundException:
                pass
        if at_menu:
            print("       At main menu.")
            return

        print(f"       Pressing Escape (attempt {attempt + 1})...")
        pyautogui.press("escape")
        time.sleep(2.0)

    print("       WARNING: Could not confirm main menu after Escape presses")


def navigate_to_train_list(train_name, extra_section_choice=None):
    """Re-navigate from main menu back to the train list (skips warning/splash).

    Stops at the train selection screen — does NOT click a specific train.
    """
    print("       Re-navigating to train list...")
    click_to_the_trains()
    click_choose_a_route()
    select_route()
    click_timetable()
    if extra_section_choice:
        select_extra_section(extra_section_choice)
    select_train(train_name)
    print("       Back at train list.")


def navigate_to_service_list(train_index, train_name, extra_section_choice=None):
    """Re-navigate from main menu back to the service list for a specific train."""
    navigate_to_train_list(train_name, extra_section_choice=extra_section_choice)
    click_train(train_index)


def scroll_service_list_to_top():
    """Scroll the service list all the way to the top.

    Scrolls up repeatedly until the list stops moving, so it works
    regardless of how far down the list was scrolled previously.
    """
    center_x = (config.SERVICE_LIST_LEFT + config.SERVICE_LIST_RIGHT) // 2
    center_y = (config.SERVICE_LIST_TOP + config.SERVICE_LIST_BOTTOM) // 2
    region = (config.SERVICE_LIST_LEFT, config.SERVICE_LIST_TOP,
              config.SERVICE_LIST_RIGHT - config.SERVICE_LIST_LEFT,
              config.SERVICE_LIST_BOTTOM - config.SERVICE_LIST_TOP)
    scroll_up = -config.SCROLL_PER_BOX * 8  # one page up

    pyautogui.moveTo(center_x, center_y)
    time.sleep(0.3)

    prev_img = np.array(pyautogui.screenshot(region=region))
    for _ in range(60):  # safety limit
        pyautogui.scroll(scroll_up)
        time.sleep(0.5)
        curr_img = np.array(pyautogui.screenshot(region=region))
        if curr_img.shape == prev_img.shape:
            diff = np.mean(np.abs(curr_img.astype(float) - prev_img.astype(float)))
            if diff < 5.0:
                break  # list stopped moving — we're at the top
        prev_img = curr_img
    time.sleep(0.5)


def scroll_service_list_down():
    """Scroll the service list down by BOXES_TO_SCROLL boxes (fixed amount)."""
    center_x = (config.SERVICE_LIST_LEFT + config.SERVICE_LIST_RIGHT) // 2
    center_y = (config.SERVICE_LIST_TOP + config.SERVICE_LIST_BOTTOM) // 2
    pyautogui.moveTo(center_x, center_y)
    time.sleep(0.3)
    pyautogui.scroll(config.SCROLL_AMOUNT)
    time.sleep(2.0)           # let scroll animation settle


def process_all_services(base_dir, train_index, train_name, max_services=None,
                         start_service=None, on_progress=None,
                         remaining_services=None, extra_section_choice=None):
    """Iterate through services using the spreadsheet, searching for each by name.

    Reads service names from the xlsx spreadsheet, types each into the search
    field to find and select it, then captures the service box screenshot,
    enters the level, captures the schedule, and uploads.

    Always returns with the game at the MAIN MENU (after exit_to_main_menu).

    Args:
        base_dir: Directory for this train's service folders (e.g. screenshots/train_01/).
        train_index: 0-based index of the current train (for re-navigation).
        train_name: Name of the current train (for re-navigation).
        max_services: Maximum services to capture (None = unlimited).
        start_service: 1-indexed service number to resume from (None = start from 1).
        on_progress: Optional callback(train_name, train_number, service_number)
                     called before each service is attempted.
        remaining_services: Optional set of service names still to find. Services not
                            in this set are skipped. Found services are removed in-place.
                            None = try all services (no filtering).
        extra_section_choice: If set, the extra section choice to select during navigation.
    """
    # Read service list from spreadsheet
    services = read_spreadsheet(train_name)
    total = len(services)
    print(f"\n       Spreadsheet: {total} services for '{train_name}'")

    batch_count = 0
    start_idx = (start_service - 1) if start_service and start_service > 1 else 0

    for svc_idx in range(start_idx, total):
        svc = services[svc_idx]
        service_number = svc_idx + 1  # 1-indexed
        svc_name = svc["name"]

        # Skip services already found on a previous train
        if remaining_services is not None and svc_name not in remaining_services:
            continue

        bound = svc["bound"]
        first_station = svc["first_station"]

        print(f"\n--- Service #{service_number}/{total}: {svc_name} ({bound}) ---")

        # Save progress so we can resume here if the bot crashes
        if on_progress:
            on_progress(train_name, train_index + 1, service_number)

        # Create per-service folder
        service_dir = os.path.join(base_dir, f"service_{service_number:03d}")
        os.makedirs(service_dir, exist_ok=True)

        # Search for the service by name
        found, y = search_and_select_service(svc_name)
        if not found:
            print(f"       SKIPPING: Could not find '{svc_name}' in search")
            shutil.rmtree(service_dir, ignore_errors=True)
            continue

        # Screenshot the selected service box
        x = (config.SERVICE_LIST_LEFT + config.SERVICE_LIST_RIGHT) // 2
        img_path = screenshot_service_box(service_dir, y, click_x=x, click_y=y)

        # Press Enter twice to load the level
        try:
            pyautogui.press("enter")
            time.sleep(1.0)
            pyautogui.press("enter")

            # Wait for level to load and get past "Get Started" screen
            wait_for_level_load(click_x=x, click_y=y, service_dir=service_dir)

            # Fetch player lat/lng for the first timetable entry
            first_lat, first_lng, _current_service = get_player_location()

            # Capture the schedule
            capture_schedule(service_dir)

            # Upload to parent app for OCR + database save
            # All 3 images (service, level info, schedule) are sent to /api/extract
            upload_ok = False
            try:
                result = upload_service(
                    service_dir, train_name,
                    train_index + 1, service_number,
                    first_station=first_station,
                    bound=bound,
                    service=svc_name,
                    first_lat=first_lat, first_lng=first_lng,
                )
                uploaded_name = result.get("service_name", "")
                if result["duplicate"]:
                    print(f"       {uploaded_name} (duplicate)")
                    upload_ok = True
                elif result["success"]:
                    print(f"       {uploaded_name}")
                    upload_ok = True
                else:
                    print(f"       Upload error: {result['error']}")
            except Exception as e:
                print(f"       Upload error: {e}")

            # Only remove from remaining once the DB confirms it's saved
            if upload_ok and remaining_services is not None:
                remaining_services.discard(svc_name)

            # Exit back to main menu
            exit_to_main_menu()

            # All services found — no need to continue
            if remaining_services is not None and not remaining_services:
                print(f"\n       All services found! Skipping rest of list.")
                break

        except TimeoutError as e:
            raise ServiceError(
                train_name, train_index + 1, service_number, e
            ) from e

        # Check service limit
        if max_services is not None and service_number >= max_services:
            print(f"\n       Reached service limit ({max_services}), stopping.")
            print(f"\nProcessed {service_number} services for this train.")
            return service_number  # at main menu

        # Check batch limit — restart game periodically to avoid degradation
        if config.BATCH_SIZE is not None:
            batch_count += 1
            if batch_count >= config.BATCH_SIZE:
                print(f"\n       Batch limit ({config.BATCH_SIZE}) reached, restarting...")
                _return_to_main_menu_from_menus()
                raise BatchRestart(
                    train_name, train_index + 1, service_number + 1
                )

        # Re-navigate to the service list for the next service
        try:
            navigate_to_service_list(train_index, train_name,
                                     extra_section_choice=extra_section_choice)
        except TimeoutError as e:
            raise ServiceError(
                train_name, train_index + 1, service_number, e
            ) from e

    # All services from spreadsheet processed
    print(f"\n       All {total} services processed. Returning to main menu...")
    _return_to_main_menu_from_menus()
    print(f"\nProcessed {total} services for this train.")
    return total


def process_all_trains(train_name, start_train=None, start_service=None,
                       on_progress=None, extra_section_choice=None):
    """Outer loop: iterate through all trains for a given train, processing services for each.

    Args:
        train_name: Name of the train to process.
        start_train: 1-indexed train number to resume from (None = start from 1).
        start_service: 1-indexed service number to resume from on the first train
                       (None = start from 1). Only applies to the start_train.
        on_progress: Optional callback(train_name, train_number, service_number)
                     called before each service is attempted.
    """
    from datetime import datetime, timedelta

    run_start = time.time()

    # Create route/train folder structure
    route_dir = os.path.join(config.SCREENSHOTS_DIR, config.ROUTE_NAME)
    if extra_section_choice:
        route_dir = os.path.join(route_dir, extra_section_choice)
    train_base_dir = os.path.join(route_dir, train_name)
    os.makedirs(train_base_dir, exist_ok=True)

    # Read spreadsheet once and build a set of service names still to find.
    # Then remove any services already recorded in the database so we never
    # waste time searching for them in-game.
    services = read_spreadsheet(train_name)
    all_service_names = {svc["name"] for svc in services}
    print(f"\n       Spreadsheet: {len(all_service_names)} unique services")

    already_recorded = get_existing_services(train_name)
    remaining_services = all_service_names - already_recorded
    if already_recorded & all_service_names:
        print(f"       Already in DB: {len(already_recorded & all_service_names)}")
    print(f"       Still to find: {len(remaining_services)}")

    if not remaining_services:
        print(f"\n       All services already recorded — skipping train '{train_name}'.")
        _return_to_main_menu_from_menus()
        return

    train_count = count_trains()
    if config.MAX_TRAINS_PER_CLASS is not None:
        train_count = min(train_count, config.MAX_TRAINS_PER_CLASS)
    print(f"\n=== Processing {train_count} trains for '{train_name}' ===\n")

    # Determine which train to start from (0-indexed)
    first_train = (start_train - 1) if start_train else 0
    if first_train > 0:
        print(f"Resuming from train {start_train}, service {start_service or 1}\n")

    train_results = []  # list of (train_number, service_count, duration_seconds)

    for train_idx in range(first_train, train_count):
        # All services already found — skip remaining trains
        if not remaining_services:
            print(f"\n       All services found! Skipping remaining trains.")
            break

        train_start = time.time()
        print(f"\n{'='*50}")
        print(f"=== Train {train_idx + 1}/{train_count} "
              f"({len(remaining_services)} services remaining) ===")
        print(f"{'='*50}")

        # Create per-train folder inside train base folder
        train_dir = os.path.join(train_base_dir, f"train_{train_idx + 1:02d}")
        os.makedirs(train_dir, exist_ok=True)

        # Click the train
        try:
            click_train(train_idx)
        except TimeoutError as e:
            raise ServiceError(train_name, train_idx + 1, 1, e) from e

        # Only use start_service for the first train in a resume
        svc_start = start_service if train_idx == first_train else None

        # Process services for this train (only tries services still in remaining_services)
        svc_count = process_all_services(
            base_dir=train_dir,
            train_index=train_idx,
            train_name=train_name,
            max_services=config.MAX_SERVICES_PER_TRAIN,
            start_service=svc_start,
            on_progress=on_progress,
            remaining_services=remaining_services,
            extra_section_choice=extra_section_choice,
        )

        train_duration = time.time() - train_start
        train_results.append((train_idx + 1, svc_count, train_duration))

        # Exit and relaunch game between trains to avoid memory issues.
        # process_all_services returns at main menu.
        # Skip relaunch if all services are already found (next iteration will break).
        if train_idx < train_count - 1 and remaining_services:
            exit_game()
            relaunch_and_navigate(train_name, extra_section_choice=extra_section_choice)

    total_duration = time.time() - run_start
    total_services = sum(r[1] for r in train_results)

    print(f"\n=== All {train_count} trains processed for '{train_name}'! ===")

    # Write summary report
    report_path = os.path.join(train_base_dir, "report.txt")
    started = datetime.fromtimestamp(run_start)
    with open(report_path, "w") as f:
        f.write(f"TSW Timetable Bot — Run Report\n")
        f.write(f"{'='*40}\n\n")
        f.write(f"Route:       {config.ROUTE_NAME}\n")
        f.write(f"Train:       {train_name}\n")
        f.write(f"Started:     {started:%Y-%m-%d %H:%M:%S}\n")
        f.write(f"Duration:    {timedelta(seconds=int(total_duration))}\n")
        f.write(f"Trains:      {train_count}\n")
        f.write(f"Services:    {total_services}\n\n")
        f.write(f"{'Train':<10} {'Services':<10} {'Duration'}\n")
        f.write(f"{'-'*10} {'-'*10} {'-'*10}\n")
        for train_num, svc_count, dur in train_results:
            f.write(f"Train {train_num:<4} {svc_count:<10} {timedelta(seconds=int(dur))}\n")

    print(f"\nReport saved to: {report_path}")

    # Write out any services that were never found on any train
    if remaining_services:
        not_found_path = os.path.join(train_base_dir, "services_not_found.txt")
        with open(not_found_path, "w") as f:
            for svc_name in sorted(remaining_services):
                f.write(f"{svc_name}\n")
        print(f"       {len(remaining_services)} services not found — "
              f"see {not_found_path}")
